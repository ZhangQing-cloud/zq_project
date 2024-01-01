from openpyxl import load_workbook,Workbook
class ExcelUtil:
    # 功能：把第一行作为表头，用作key 关键字，生成字典对象，存入列表返回
    @staticmethod
    def readDict(file):
        wb = load_workbook(file)
        sheet = wb.active
        header = [x.value for x in sheet[1]]
        wb_rows_values = sheet.values

        lst = []    # 列表
        row_index = 0
        for row in wb_rows_values:
            row_index += 1
            if row_index == 1:
                continue

            col_index = 0
            row_data = {}
            for v in row:
                row_data[header[col_index]] = v
                col_index += 1
            lst.append(row_data)
        return lst

    # 封装方法：把列表数据直接导出到指定的excel文件中
    @staticmethod
    def rowToExcel(dataList,file):
        wb = Workbook()
        sheet = wb.active # 工作表
        sheet.append(dataList)
        wb.save(file)

    @staticmethod
    def listToExcel(dictList: list, file):
        wb = Workbook()
        sheet = wb.active  # 工作表

        for dict in dictList:
            print(dict)
            row = list(dict.values())
            sheet.append(row)
        wb.save(file)

